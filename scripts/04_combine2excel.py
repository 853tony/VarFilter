import os
import argparse
import pandas as pd
from openpyxl import Workbook

def main(input_folder, output_prefix, group_pattern):
    # Create a dictionary to store the workbooks for each group
    workbooks = {pattern: Workbook() for pattern in group_patterns}

    # Iterate over all the files in the folder
    for filename in os.listdir(input_folder):

        # Check if the file is a TSV file
        if filename.endswith(".tsv"):
            # Check if the group pattern is present in the filename
            for pattern in group_patterns:
                if pattern in filename:
                    group_name = pattern
                    break
            else:
                # If no group pattern is found, skip the file
                continue
            
            # Get the workbook for the current group
            workbook = workbooks[group_name]
            
            # Create the full path to the TSV file
            file_path = os.path.join(input_folder, filename)
            
            # Read the TSV file into a DataFrame
            df = pd.read_csv(file_path, sep="\t", na_filter=False, keep_default_na=False)
            
            # Sort the DataFrame based on the values in the second and first columns
            df = df.sort_values(by=[df.columns[0], df.columns[1]])
            
            # Create a new worksheet using the TSV file name as the worksheet name
            sheet_name = os.path.splitext(filename)[0]
            sheet = workbook.create_sheet(title=sheet_name)
            
            # Write the header to the worksheet
            for col_idx, column in enumerate(df.columns, start=1):
                sheet.cell(row=1, column=col_idx).value = column
            
            # Write the DataFrame to the worksheet
            for r in range(df.shape[0]):
                for c in range(df.shape[1]):
                    value = df.iloc[r,c]
                    sheet.cell(row=r+2, column=c+1).value = "" if pd.isna(value) else str(value)
            
            print(f"Written {filename} to worksheet {sheet_name} in group {group_name}")

    # Save each workbook to a separate Excel file
    for group_name, workbook in workbooks.items():
        output_file = f"{output_prefix}_{group_name}.xlsx"
        workbook.remove(workbook["Sheet"])  # Remove the default empty worksheet
        workbook.save(output_file)
        print(f"Integrated files for group {group_name} into {output_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Integrate TSV files into Excel workbooks based on groups")
    parser.add_argument("input_folder", help="Path to the folder containing the TSV files")
    parser.add_argument("output_prefix", help="Prefix for the output Excel files")
    parser.add_argument("group_patterns", help="Comma-separated list of patterns to use for grouping the files")
    args = parser.parse_args()

    group_patterns = args.group_patterns.split(",")
    main(args.input_folder, args.output_prefix, args.group_patterns)
